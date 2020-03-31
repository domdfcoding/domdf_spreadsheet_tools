#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  domdf_spreadsheet_tools.py
"""Tools for creating and formatting spreadsheets with Python and OpenPyXL"""
#
#  Copyright 2018-2019 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU Lesser General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU Lesser General Public License for more details.
#
#  You should have received a copy of the GNU Lesser General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#

# stdlib
import os
import io
import csv
import locale
import traceback

# 3rd party
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from domdf_python_tools.utils import as_text


__author__ = "Dominic Davis-Foster"
__copyright__ = "Copyright 2018-2019 Dominic Davis-Foster"

__license__ = "LGPL"
__version__ = "0.1.2"
__email__ = "dominic@davis-foster.co.uk"


def append_to_xlsx(
		csv_input_file, xlsx_output_file, sheet_title=None, separator=",",
		overwrite=False, use_io=False, toFloats=False
):
	"""
	Add CSV file to xlsx file as a new worksheet
	
	:param csv_input_file: filepath of CSV file to
	:type csv_input_file: str
	:param xlsx_output_file: filepath of xlsx file
	:type xlsx_output_file: str
	:param sheet_title: Title of sheet to append. Default is name of csv_input_file
	:type sheet_title: str
	:param separator: Separator for reading CSV file, default to comma
	:type separator: str
	:param overwrite: Whether to overwrite the xlsx output file (i.e. create a new file containing just the new sheet)
	:type overwrite: bool
	:param use_io: Whether to use the io module
	:type use_io: bool
	:param toFloats: Whether to read strings with thousand separators as floats
	:type toFloats: bool
	"""
	
	# Setup for reading strings with thousand separators as floats
	# From https://stackoverflow.com/a/31074271
	locale.setlocale(locale.LC_ALL, "")
	
	if sheet_title is None:
		sheet_title = os.path.splitext(os.path.basename(csv_input_file))[0]
	
	if overwrite:
		wb = Workbook()
		ws = wb.active
		wb.remove_sheet(ws)
	else:
		wb = load_workbook(xlsx_output_file)
	
	wb.create_sheet(sheet_title)
	ws = wb[sheet_title]
	
	if use_io:
		f = io.open(csv_input_file, encoding='latin-1')
	else:
		f = open(csv_input_file)
	reader = csv.reader(f, delimiter=separator)
	
	for row in reader:
		try:
			if toFloats:
				row_buffer = []
				for cell in row:
					try:
						row_buffer.append(locale.atof(cell))
					except:
						row_buffer.append(cell)
				ws.append(row_buffer)
			else:
				ws.append(row)
		except:
			traceback.print_exc()  # print the error
			print(row)
	f.close()
	
	wb.save(xlsx_output_file)


def format_sheet(ws, number_format_list=None, width_list=None, alignment_list=None):
	"""
	Format columns of an xlsx worksheet
	
	:param ws: The worksheet to format
	:type ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
	:param number_format_list: dictionary of number format strings for each column letter
	:type number_format_list: dict
	:param width_list: dictionary of widths for each column letter
	:type width_list: dict
	:param alignment_list: dictionary of alignments (left, right, center) for each column letter
	:type alignment_list: dict
	"""
	
	# for row in ws.iter_rows("A1:{}{}".format(get_column_letter(ws.max_column), ws.max_row)):
	for row in ws["A1:{}{}".format(get_column_letter(ws.max_column), ws.max_row)]:
		for cell in row:
			cell.alignment = Alignment(vertical="center", wrap_text=False)
	
	if number_format_list:
		for column in number_format_list:
			# for row in ws.iter_rows('{0}{1}:{0}{2}'.format(column, 3, ws.max_row)):
			for row in ws['{0}{1}:{0}{2}'.format(column, 3, ws.max_row)]:
				for cell in row:
					cell.number_format = number_format_list[column]
	
	for column_cells in ws.columns:
		length = max(len(as_text(cell.value)) for cell in column_cells)
		if length < 1.0:
			length = 1.0
		ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
	# ws.column_dimensions[column_cells[0].column].bestFit = True
	
	if width_list:
		for column in width_list:
			if width_list[column] == 0:
				ws.column_dimensions[column].hidden = True
			else:
				ws.column_dimensions[column].width = width_list[column]
	
	if alignment_list:
		for column in alignment_list:
			# for row in ws.iter_rows("{0}{1}:{0}{2}".format(column, ws.min_row, ws.max_row)):
			for row in ws["{0}{1}:{0}{2}".format(column, ws.min_row, ws.max_row)]:
				for cell in row:
					cell.alignment = Alignment(
						horizontal=alignment_list[column],
						vertical="center",
						wrap_text=False
					)
	

def format_header(ws, alignment_list, start_row=1, end_row=1):
	"""
	Format the alignment of the header rows of a worksheet
	
	:param ws: The worksheet to format
	:type ws: :class:`openpyxl.worksheet.worksheet.Worksheet`
	:param alignment_list: dictionary of alignments (left, right, center) for each column letter
	:type alignment_list: dict
	:param start_row: The row to start formatting on
	:type start_row: int
	:param end_row: The row to end formatting on
	:type end_row: int
	"""
	
	for column in alignment_list:
		# for row in ws.iter_rows("{0}{1}:{0}{2}".format(column, start_row, end_row)):
		for row in ws["{0}{1}:{0}{2}".format(column, start_row, end_row)]:
			for cell in row:
				cell.alignment = Alignment(
					horizontal=alignment_list[column],
					vertical="center",
					wrap_text=False
				)


def make_column_property_dict(indict, outdict=None, offset_dict=None, repeat=1, length=1):
	"""
	Generate property lists from integer values
	
	:param indict: Property values to add to the property dict
	:type indict: dict
	:param outdict: Dictionary of properties for each column letter
	:type outdict: dict
	:param offset_dict:
	:type offset_dict:
	:param repeat:
	:type repeat:
	:param length:
	:type length:
	TODO: Finish this docstring; check usage in GunShotMatch
	:return:
	"""

	if not outdict:
		outdict = {}
	for index in indict:
		for offset in range(repeat):
			outdict[get_column_letter(int(index) + (length * offset))] = indict[index]
	
	if offset_dict:
		offset = repeat * length
		for index in offset_dict:
			outdict[get_column_letter(int(index) + offset)] = offset_dict[index]
	
	return outdict